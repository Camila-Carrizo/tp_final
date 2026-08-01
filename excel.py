"""
excel.py
--------
Escribe cada fila de la simulación al .xlsx (sin acumular en memoria)
y permite leerlas después (p. ej. desde la UI).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Orden de columnas (planilla) — grupos = mismo color de encabezado
COLUMNAS = [
    "EVENTO",
    "RELOJ",
    "RND_LLEGADA_ASCENSOR",
    "LLEGADA_ASCENSOR",
    "PROXIMA_LLEGADA_ASCENSOR",
    "RND_H",
    "H",
    "RND_P",
    "P",
    "RND_LLEGADA_PASAJERO",
    "LLEGADA_PASAJERO",
    "PROXIMA_LLEGADA_PASAJERO",
    "RND_DIRECCION_PASAJERO",
    "DIRECCION_PASAJERO",
    "DIRECCION_ASCENSOR",
    "ESTADO_ASCENSOR",
    "ESPACIO_DISPONIBLE",
    "INICIO_DETENCION",
    "FIN_DESCENSO",
    "FIN_ASCENSO",
    "FIN_ESPERA",
    "COLA_BAJA",
    "COLA_SUBE",
    "ACUMULADOR_PERMANENCIA",
]

# Título visible en Excel/UI (la clave interna sigue siendo la de COLUMNAS)
TITULOS_COLUMNA = {
    "RELOJ": "RELOJ (seg)",
}


def titulo_columna(clave: str) -> str:
    return TITULOS_COLUMNA.get(clave, clave)


def caracteres_columna(clave: str) -> int:
    """Ancho en caracteres para que el encabezado se vea completo."""
    return max(len(titulo_columna(clave)), 10)


# Color por columna (encabezado). Mismo hex = mismo grupo.
COLORES_ENCABEZADO = {
    "EVENTO": "D9D2E9",  
    "RELOJ": "CFE2F3",  
    "RND_LLEGADA_ASCENSOR": "D9EAD3",  
    "LLEGADA_ASCENSOR": "D9EAD3",
    "PROXIMA_LLEGADA_ASCENSOR": "D9EAD3",
    "RND_H": "FCE5CD",  
    "H": "FCE5CD",
    "RND_P": "E6B8AF",  
    "P": "E6B8AF",
    "RND_LLEGADA_PASAJERO": "D0E0E3",  
    "LLEGADA_PASAJERO": "D0E0E3",
    "PROXIMA_LLEGADA_PASAJERO": "D0E0E3",
    "RND_DIRECCION_PASAJERO": "D0E0E3",
    "DIRECCION_PASAJERO": "D0E0E3",
    "DIRECCION_ASCENSOR": "EAD1DC",  
    "ESTADO_ASCENSOR": "EAD1DC",
    "ESPACIO_DISPONIBLE": "EAD1DC",
    "INICIO_DETENCION": "EAD1DC",
    "FIN_DESCENSO": "F4CCCC",  
    "FIN_ASCENSO": "FFF2CC",  
    "FIN_ESPERA": "B6D7A8",  
    "COLA_BAJA": "B4C6E7",  
    "COLA_SUBE": "B4C6E7",
    "ACUMULADOR_PERMANENCIA": "EFEFEF",  
}


_COLORES_TEXTO_OSCURO = set(COLORES_ENCABEZADO.values())


class EscritorExcel:
    """Libro en disco: encabezado + append de filas + guardar al final."""

    def __init__(self, ruta_archivo: str | Path):
        self.ruta = Path(ruta_archivo)
        # Cada corrida arranca de cero: borra el Excel anterior si existe
        if self.ruta.exists():
            self.ruta.unlink()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Simulacion"
        self.ws.append([titulo_columna(c) for c in COLUMNAS])
        self._estilo_encabezados()
        self._ajustar_anchos()

    def _estilo_encabezados(self) -> None:
        for idx, nombre in enumerate(COLUMNAS, start=1):
            celda = self.ws.cell(row=1, column=idx)
            color = COLORES_ENCABEZADO.get(nombre, "1F4E79")
            celda.fill = PatternFill("solid", fgColor=color)
            texto = "000000" if color in _COLORES_TEXTO_OSCURO else "FFFFFF"
            celda.font = Font(color=texto, bold=True)
            celda.alignment = Alignment(horizontal="center", wrap_text=False)

    def _ajustar_anchos(self) -> None:
        """Cada columna tan ancha como su encabezado (+ margen)."""
        for idx, nombre in enumerate(COLUMNAS, start=1):
            letra = get_column_letter(idx)
            self.ws.column_dimensions[letra].width = caracteres_columna(nombre) + 2

    def agregar_fila(self, estado: dict) -> None:
        self.ws.append([_celda(estado.get(col)) for col in COLUMNAS])

    def escribir_parametros(self, parametros: dict) -> None:
        _escribir_parametros(self.wb, parametros)

    def guardar(self) -> Path:
        self.ruta.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.ruta)
        return self.ruta


def leer_filas(ruta_archivo: str | Path) -> list[dict]:
    """Lee el .xlsx y devuelve una lista de dicts (una por fila de datos)."""
    wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        return []
    titulo_a_clave = {titulo_columna(c): c for c in COLUMNAS}
    encabezados = []
    for h in filas[0]:
        texto = str(h) if h is not None else ""
        encabezados.append(titulo_a_clave.get(texto, texto))
    return [dict(zip(encabezados, fila)) for fila in filas[1:]]


def exportar_resultados(
    filas: list[dict],
    ruta_archivo: str,
    parametros: dict | None = None,
) -> Path:
    """
    Alternativa: recibe ya todas las filas y las guarda de una.
    La simulación normal usa EscritorExcel (append por evento).
    """
    escritor = EscritorExcel(ruta_archivo)
    for estado in filas:
        escritor.agregar_fila(estado)
    if parametros:
        _escribir_parametros(escritor.wb, parametros)
    return escritor.guardar()


def _escribir_parametros(wb: Workbook, parametros: dict) -> None:
    if "Parametros" in wb.sheetnames:
        ws = wb["Parametros"]
    else:
        ws = wb.create_sheet("Parametros")
    ws.delete_rows(1, ws.max_row)
    ws.append(["clave", "valor"])
    for clave, valor in parametros.items():
        ws.append([clave, _celda(valor)])


def _celda(valor):
    """None → celda vacía."""
    return "" if valor is None else valor
